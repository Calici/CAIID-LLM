from __future__ import annotations

import asyncio
import csv
import pathlib
import subprocess
import xml.etree.ElementTree as ET
from typing import Protocol, final

import openpyxl
from docx import Document

from app.libs.expected import Expected


@final
class FileReaderError(RuntimeError):
    pass


class FileReader(Protocol):
    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]: ...


class _FileReader:
    @staticmethod
    def create_expected(v: str | FileReaderError):
        return Expected(str, FileReaderError, v)


@final
class PDFReader:
    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]:
        try:
            proc = await asyncio.create_subprocess_exec(
                "pdftotext",
                str(p),
                "-",
                stdout=asyncio.subprocess.PIPE,
                stdin=asyncio.subprocess.DEVNULL,
                stderr=asyncio.subprocess.DEVNULL,
            )
            assert proc.stdout is not None  # unreachable
            output = await proc.stdout.read()
            return _FileReader.create_expected(output.decode())
        except subprocess.CalledProcessError:
            return _FileReader.create_expected(
                FileReaderError("pdftotext conversion fail")
            )
        except UnicodeDecodeError:
            return _FileReader.create_expected(
                FileReaderError("pdftotext conversion fail")
            )


@final
class PlainTextReader:
    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]:
        try:
            f = open(p, "r", encoding="utf-8-sig")
            return _FileReader.create_expected(f.read())
        except IOError:
            return _FileReader.create_expected(FileReaderError("file read error"))


@final
class DocxReader:
    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]:
        try:
            doc = Document(str(p))
            return _FileReader.create_expected(
                "\n".join([p.text for p in doc.paragraphs])
            )
        except Exception as e:
            return _FileReader.create_expected(FileReaderError(str(e)))


@final
class ExtReader:
    def __init__(self, fallback_reader: FileReader):
        self.readers: dict[str, FileReader] = {"*": fallback_reader}

    def add_reader(self, ext: str, reader: FileReader) -> ExtReader:
        self.readers[ext] = reader
        return self

    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]:
        reader = self.readers.get(p.suffix, self.readers["*"])
        return await reader.read_file(p)


@final
class CSVReader:
    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]:
        try:
            with open(p, "r") as f:
                reader = csv.reader(f)
                headers = next(reader)
                rows = list(reader)
                payload = self.build_table_xml(headers, rows)
                return _FileReader.create_expected(payload)
        except IOError:
            return _FileReader.create_expected(FileReaderError("file read error"))
        except StopIteration:
            return _FileReader.create_expected(FileReaderError("csv header missing"))

    @staticmethod
    def build_table_xml(headers: list[str], rows: list[list[str]]) -> str:
        table_elem = ET.Element("table")
        header_elem = ET.SubElement(table_elem, "header")
        header_elem.text = ", ".join(str(value) for value in headers)
        for row in rows:
            row_elem = ET.SubElement(table_elem, "row")
            row_elem.text = ", ".join(str(cell) for cell in row)
        return ET.tostring(table_elem, encoding="unicode")


@final
class XlsxReader:
    async def read_file(self, p: pathlib.Path) -> Expected[str, FileReaderError]:
        doc = openpyxl.load_workbook(p)
        root = ET.Element("table")
        for sheet in doc.worksheets:
            sheet_xml = ET.SubElement(root, "sheet", {"name": sheet.title})
            for row in range(0, sheet.max_row):
                row_xml = ET.SubElement(sheet_xml, "row")
                row_xml.text = ", ".join(
                    str(col[row].value) for col in sheet.iter_cols(0, sheet.max_column)
                )
        return _FileReader.create_expected(ET.tostring(root, encoding="unicode"))


file_reader: FileReader = (
    ExtReader(PlainTextReader())
    .add_reader(".doc", DocxReader())
    .add_reader(".docx", DocxReader())
    .add_reader(".pdf", PDFReader())
    .add_reader(".xlsx", XlsxReader())
)
